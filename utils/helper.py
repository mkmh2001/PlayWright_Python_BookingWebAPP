import json
import os

from openpyxl import Workbook, load_workbook
from playwright.sync_api import expect
import pandas as pd

class Helper:
    result = []
    first_run = True

    def __init__(self, page):
        self.page = page

    @staticmethod
    def select_date(page, month_year: str, day: str):
        calendar_btn = page.locator("//button[@data-ui-name = 'button_date_segment_0']")
        calendar_btn.wait_for()
        calendar_btn.click()

        # keep clicking next until month visible -> //h3[contains(text(), '{month_year}')]
        while not page.locator(f"//h3[contains(text(), '{month_year}')]").first.is_visible():
            next_btn = page.locator("//button[contains(@class, 'control--next')]")
            expect(next_btn).to_be_visible()
            next_btn.click()
            page.wait_for_timeout(300)  # short wait for re-render

        # click the specific day -> //h3[contains(text(), '{month_year}')]/following-sibling::table//span[text() = '{day}']']
        day_locator = page.locator(
            f"//h3[contains(text(), '{month_year}')]/following-sibling::table//span[text() = '{day}']")
        expect(day_locator).to_be_visible()
        day_locator.click()

    @staticmethod
    def click_btn(element):
        try:
            element.wait_for()
            element.click()
        except Exception as e:
            print("Error Clicking on Btn" + e)

    @staticmethod
    def check_radio_btn(element):
        try:
            element.wait_for()
            element.click()
        except Exception as e:
            print("Error Clicking on Radio Btn" + e)

    @staticmethod
    def set_dropdown_value(element, option):
        try:
            element.wait_for()
            element.select_option(option)
        except Exception as e:
            print("Error setting the option to drop down" + e)

    @staticmethod
    def set_value_in_el(element, input_text):
        try:
            element.wait_for()
            element.click()
            element.fill(input_text)
        except Exception as e:
            print("Error setting the value " + input_text + " to element " + e)

    @staticmethod
    def save_data_to_excel(air_line_details: list, file_name="TestResults.xlsx"):
        # Ensure 'results' folder exists
        folder = "test_results"
        if not os.path.exists(folder):
            os.makedirs(folder)  # create folder if it doesn't exist

        file_path = os.path.join(folder, file_name)  # path inside results folder

        # Column headers
        headers = ["Airline","From", "To", "Price", "Departure Time", "Arrival Time"]

        if Helper.first_run:
            # First test → create new workbook (clears old content)
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(headers)  # write headers
            Helper.first_run = False
        else:
            # Subsequent tests → append to existing workbook
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Results"
                ws.append(headers)

        # Append test data
        for row in air_line_details:
            ws.append(row)

        wb.save(file_path)  # save inside results folder

    @staticmethod
    def get_test_data():
        data = pd.read_excel("data/test_data.xlsx")

        passenger_details = []
        for index, row in data.iterrows():
            passenger_details.append(row.to_dict())

        return passenger_details

    @staticmethod
    def get_location_code(location):
        with open('data/city_name_mapping.json') as json_file:
            data = json.load(json_file)
            print(data[location])
            return data[location]
